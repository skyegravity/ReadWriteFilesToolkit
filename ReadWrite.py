from openpyxl import Workbook,load_workbook
import os
import csv
import xlrd
def read_xlsx(file_name,sheet_name=None):
    file_path = os.path.normpath(file_name)
    if os.path.isfile(file_path):
        book = load_workbook(file_path,data_only=True)
        if sheet_name:
           sheet = book.get_sheet_by_name(sheet_name)
        else:
           sheet = book.worksheets[0]
        nrow = sheet.max_row
        ncol = sheet.max_column
        rows = []
        for row in sheet.rows:
            row1 = [e.value for e in row]
            rows.append(row1)
        return rows
    else:
        print('Excel File %s does not exist!'% file_name)

def read_xls(file_name,sheet_name=None):
    file_path = os.path.normpath(file_name)
    book = xlrd.open_workbook(file_path)
    if sheet_name:
       sheet = book.sheet_by_name(sheet_name)
    else:
       sheet = book.sheet_by_index(0)
    rows = []
    for row_num in range(sheet.nrows):
        row_value = sheet.row_values(row_num)
        rows.append(row_value)
    return rows


def write_xlsx(file_name,input):
    wb = Workbook()
    ws = wb.active
    for e in input:
       ws.append(e)
    wb.save(os.path.normpath(file_name))

def write_txt(file_name,input):
    f = open(file_name, 'wb')
    for e in input:
       f.write(e+'\r\n')  # python will convert \n to os.linesep
    f.close()

def read_csv(file_name):
    file_path = os.path.normpath(file_name)
    with open(file_path,'rb') as f:
        reader = csv.reader(f)
        rows =[]
        for row in reader:
            rows.append(row)
    return rows

def write_csv(file_name,file):
    with open(file_name, 'wb') as csvfile:
        #writer = csv.writer(csvfile, delimiter=' ',quotechar='|', quoting=csv.QUOTE_MINIMAL)
        writer = csv.writer(csvfile)
        for row in file:
            writer.writerow(row)
